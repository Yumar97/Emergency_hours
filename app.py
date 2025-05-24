from flask import Flask, render_template, request, redirect, url_for, flash, session, send_file, Blueprint
from flask_sqlalchemy import SQLAlchemy
from datetime import datetime, date, timedelta
from werkzeug.security import generate_password_hash, check_password_hash
from sqlalchemy import and_, or_, func
import calendar
import colorsys
import pandas as pd
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.styles.fills import GradientFill
from openpyxl import Workbook
import flask as get_column_letter
from sqlalchemy.sql import func
from openpyxl.utils import get_column_letter






app = Flask(__name__)
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///database.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.secret_key = 'your_secret_key'
db = SQLAlchemy(app)

# Función auxiliar para capitalizar nombres
def capitalizar_nombre(nombre):
    """Convierte un nombre completo a formato con primera letra de cada palabra en mayúscula."""
    return ' '.join(word.capitalize() for word in nombre.split())

# Filtro de Jinja2 para capitalizar nombres
@app.template_filter('capitalizar')
def capitalizar_filter(nombre):
    """Filtro para capitalizar la primera letra de cada palabra en Jinja2."""
    return ' '.join(word.capitalize() for word in nombre.split())

# Modelos
class User(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(50), nullable=False, unique=True)
    password_hash = db.Column(db.String(100), nullable=False)

    def set_password(self, password):
        self.password_hash = generate_password_hash(password)

    def check_password(self, password):
        return check_password_hash(self.password_hash, password)

class Turno(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    doctor = db.Column(db.String(100), nullable=False)
    area = db.Column(db.String(100), nullable=False)
    fecha_trabajo = db.Column(db.Date, nullable=False)
    horas_trabajadas = db.Column(db.Integer, nullable=False, default=6)
    turno = db.Column(db.String(10), nullable=False)
    tipo_horas = db.Column(db.String(10), nullable=False)

class Doctor(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False, unique=True)

# Listas de áreas
opciones_diurnas = [
    "Observ - Medico 1 (Jg)", "Observ - Medico 2", "Shock T", "Observ - Medico 3",
    "Tópico 1", "Tópico 2", "Tópico 3", "Tópico 4", "Triaje 1", "Triaje 2"
]

opciones_nocturnas = ["Jefe G.", "G.N.", "Triaje"]

@app.route('/')
def index():
    if 'user_id' not in session:
        return redirect(url_for('login'))

    doctores = Doctor.query.all()
    print("Doctores:", [(d.id, d.name) for d in doctores])  # Debugging print statement
    doctor_filtro = request.args.getlist('doctor_filtro', type=str)
    mes_filtro = request.args.getlist('mes_filtro', type=int)
    scroll_position = request.args.get('scroll_position', 0, type=int)
    año_actual = datetime.now().year

    query = Turno.query

    if doctor_filtro:
        if 'todos' in doctor_filtro:
            if len(doctor_filtro) > 1:
                flash("No se puede combinar 'Todos los doctores' con selecciones individuales.", "warning")
                return redirect(url_for('index'))
            query = query
        else:
            query = query.filter(Turno.doctor.in_(doctor_filtro))

    if mes_filtro:
        conditions = []
        for mes in mes_filtro:
            primer_dia = date(año_actual, mes, 1)
            ultimo_dia = date(año_actual, mes, calendar.monthrange(año_actual, mes)[1])
            conditions.append(Turno.fecha_trabajo.between(primer_dia, ultimo_dia))
        if conditions:
            query = query.filter(or_(*conditions))

    turnos = query.all()

    return render_template(
        'index.html',
        username=session['username'],
        turnos=turnos,
        opciones_diurnas=opciones_diurnas,
        opciones_nocturnas=opciones_nocturnas,
        datetime=datetime,
        doctores=doctores,
        doctor_filtro=doctor_filtro,
        mes_filtro=[str(m) for m in mes_filtro],
        scroll_position=scroll_position
    )

@app.route('/nuevo_doctor', methods=['GET', 'POST'])
def nuevo_doctor():
    if 'user_id' not in session:
        flash('Debe iniciar sesión primero', 'danger')
        return redirect(url_for('login'))

    if request.method == 'POST':
        nombre = request.form.get('nombre')
        apellido = request.form.get('apellido')
        
        # Check if both fields are provided
        if nombre and apellido:
            # Combine nombre and apellido into a single name
            nombre_doctor = f"{nombre.strip()} {apellido.strip()}"
            # Capitalizar el nombre completo
            nombre_doctor = capitalizar_nombre(nombre_doctor)
            doctor_existente = Doctor.query.filter_by(name=nombre_doctor).first()
            if doctor_existente:
                flash("❌ Error: El doctor ya está registrado.", "danger")
            else:
                try:
                    nuevo_doctor = Doctor(name=nombre_doctor)
                    db.session.add(nuevo_doctor)
                    db.session.commit()
                    flash("✅ Doctor agregado con éxito.", "success")
                except Exception as e:
                    db.session.rollback()
                    flash(f"❌ Error: Hubo un problema al registrar el doctor. {str(e)}", "danger")
        else:
            flash("⚠️ Tanto el nombre como el apellido del doctor son obligatorios.", "warning")
    
    doctores = Doctor.query.all()
    scroll_position = request.args.get('scroll_position', 0, type=int)
    return render_template('nuevo_doctor.html', doctores=doctores, scroll_position=scroll_position)

@app.route('/eliminar_doctor/<int:id>', methods=['POST'])
def eliminar_doctor(id):
    if 'user_id' not in session:
        flash('Debe iniciar sesión primero', 'danger')
        return redirect(url_for('login'))

    scroll_position = request.form.get('scroll_position', 0, type=int)
    doctor = Doctor.query.get(id)
    if doctor:
        try:
            db.session.delete(doctor)
            db.session.commit()
            flash("✅ Doctor eliminado con éxito.", "success")
        except:
            db.session.rollback()
            flash("❌ Error al eliminar el doctor.", "danger")
    else:
        flash("⚠️ Doctor no encontrado.", "warning")
    return redirect(url_for('nuevo_doctor', scroll_position=scroll_position))

@app.route('/editar_doctor/<int:id>', methods=['GET', 'POST'])
def editar_doctor(id):
    if 'user_id' not in session:
        flash('Debe iniciar sesión primero', 'danger')
        return redirect(url_for('login'))

    doctor = Doctor.query.get_or_404(id)
    scroll_position = request.args.get('scroll_position', 0, type=int)

    if request.method == 'POST':
        scroll_position = request.form.get('scroll_position', 0, type=int)
        nombre = request.form.get('nombre')

        if not nombre:
            flash('El nombre del doctor es obligatorio.', 'danger')
            return render_template('editar_doctor.html', doctor=doctor, scroll_position=scroll_position)

        # Capitalizar el nombre
        nombre = capitalizar_nombre(nombre)

        # Check if the new name already exists (excluding the current doctor)
        existing_doctor = Doctor.query.filter_by(name=nombre).first()
        if existing_doctor and existing_doctor.id != doctor.id:
            flash('El nombre del doctor ya está registrado.', 'danger')
            return render_template('editar_doctor.html', doctor=doctor, scroll_position=scroll_position)

        # Update the doctor's name
        doctor.name = nombre

        try:
            db.session.commit()
            flash('Doctor actualizado correctamente.', 'success')
            return redirect(url_for('nuevo_doctor', scroll_position=scroll_position))
        except Exception as e:
            db.session.rollback()
            flash('Error al actualizar el doctor: ' + str(e), 'danger')

    return render_template('editar_doctor.html', doctor=doctor, scroll_position=scroll_position)

@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']

        if User.query.filter_by(username=username).first():
            flash('El usuario ya existe. Elija otro.', 'danger')
            return redirect(url_for('register'))

        new_user = User(username=username)
        new_user.set_password(password)

        db.session.add(new_user)
        db.session.commit()

        flash('Registro exitoso. Ahora puede iniciar sesión.', 'success')
        return redirect(url_for('login'))

    return render_template('register.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']

        user = User.query.filter_by(username=username).first()
        if user and user.check_password(password):
            session['user_id'] = user.id
            session['username'] = user.username
            flash('Inicio de sesión exitoso.', 'success')
            return redirect(url_for('index'))
        else:
            flash('Usuario o contraseña incorrectos.', 'danger')

    return render_template('login.html')

@app.route('/logout')
def logout():
    session.pop('user_id', None)
    session.pop('username', None)
    flash('Sesión cerrada.', 'success')
    return redirect(url_for('login'))

turnos_bp = Blueprint('turnos', __name__)

@app.route('/turnos', methods=['GET', 'POST'])
def turnos():
    if 'user_id' not in session:
        flash('Debe iniciar sesión primero', 'danger')
        return redirect(url_for('login'))

    if request.method == "POST":
        nombre = request.form.get("nombre")
        area = request.form.get("area")
        fecha_trabajo = request.form.get("fecha_trabajo")
        horas_trabajadas = int(request.form.get("horas_trabajadas"))
        turno = request.form.get("turno").upper()
        horas_normales = request.form.get("horas_normales")
        horas_extras = request.form.get("horas_extras")
        scroll_position = request.form.get('scroll_position', 0, type=int)

        # Validate exactly one type of hours is selected
        if (horas_normales and horas_extras) or (not horas_normales and not horas_extras):
            flash("Debe seleccionar exactamente un tipo de horas: Normales o Extras.", "danger")
            return redirect(url_for("index", scroll_position=scroll_position))

        tipo_horas = "Normal" if horas_normales else "Extra"

        # Validate required fields
        if not all([nombre, area, fecha_trabajo, horas_trabajadas, turno]):
            flash("Todos los campos son obligatorios.", "danger")
            return redirect(url_for("index", scroll_position=scroll_position))

        # Validate area compatibility with turn
        if turno == "NOCHE" and area not in opciones_nocturnas:
            flash("Área inválida para el turno Noche.", "danger")
            return redirect(url_for("index", scroll_position=scroll_position))
        elif turno != "NOCHE" and area not in opciones_diurnas:
            flash("Área inválida para el turno seleccionado.", "danger")
            return redirect(url_for("index", scroll_position=scroll_position))

        try:
            fecha = datetime.strptime(fecha_trabajo, "%Y-%m-%d").date()
            inicio_mes = date(fecha.year, fecha.month, 1)
            fin_mes = date(fecha.year, fecha.month, calendar.monthrange(fecha.year, fecha.month)[1])

            # Capitalizar el nombre
            nombre_capitalizado = capitalizar_nombre(nombre)

            # *** NEW VALIDATION: Check for existing shift with same date, turn, and area ***
            existing_turno_conflict = Turno.query.filter_by(
                fecha_trabajo=fecha,
                turno=turno,
                area=area
            ).first()
            if existing_turno_conflict:
                flash(f"Ya existe un turno registrado para el área {area} en el turno {turno} del día {fecha.strftime('%d/%m/%Y')}.", "danger")
                return redirect(url_for("index", scroll_position=scroll_position))

            # Check for existing shift to prevent duplicates
            existing_turno = Turno.query.filter_by(
                doctor=nombre_capitalizado,
                fecha_trabajo=fecha,
                turno=turno,
                area=area,
                tipo_horas=tipo_horas,
                horas_trabajadas=horas_trabajadas
            ).first()

            if existing_turno:
                flash("Este turno ya está registrado.", "warning")
                return redirect(url_for("index", scroll_position=scroll_position))

            # Calculate current monthly hours
            total_horas_normales = db.session.query(func.sum(Turno.horas_trabajadas)).filter(
                and_(
                    func.lower(Turno.doctor) == nombre.lower(),
                    Turno.fecha_trabajo.between(inicio_mes, fin_mes),
                    Turno.tipo_horas == "Normal"
                )
            ).scalar() or 0

            total_horas_extras = db.session.query(func.sum(Turno.horas_trabajadas)).filter(
                and_(
                    func.lower(Turno.doctor) == nombre.lower(),
                    Turno.fecha_trabajo.between(inicio_mes, fin_mes),
                    Turno.tipo_horas == "Extra"
                )
            ).scalar() or 0

            # Check monthly hour limits
            if tipo_horas == "Normal" and total_horas_normales + horas_trabajadas > 150:
                flash(f"No se pueden asignar más de 150 horas normales por mes. Horas actuales: {total_horas_normales}", "danger")
                return redirect(url_for("index", scroll_position=scroll_position))

            if tipo_horas == "Extra" and total_horas_extras + horas_trabajadas > 100:
                flash(f"No se pueden asignar más de 100 horas extras por mes. Horas actuales: {total_horas_extras}", "danger")
                return redirect(url_for("index", scroll_position=scroll_position))

            # Create and save new shift
            nuevo_turno = Turno(
                doctor=nombre_capitalizado,
                area=area,
                fecha_trabajo=fecha,
                horas_trabajadas=horas_trabajadas,
                turno=turno,
                tipo_horas=tipo_horas
            )

            db.session.add(nuevo_turno)
            db.session.commit()
            flash("Turno agregado con éxito.", "success")
        except Exception as e:
            db.session.rollback()
            flash(f"Error al guardar el turno: {str(e)}", "danger")

        return redirect(url_for("index", scroll_position=scroll_position))

    # GET request: Redirect to index to avoid duplicating logic
    return redirect(url_for('index'))

@app.route('/borrar_turno/<int:id>', methods=['POST'])
def borrar_turno(id):
    if 'user_id' not in session:
        flash('Debe iniciar sesión primero', 'danger')
        return redirect(url_for('login'))

    turno = Turno.query.get(id)
    if turno:
        try:
            db.session.delete(turno)
            db.session.commit()
            flash('Turno eliminado correctamente', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'Error al eliminar turno: {str(e)}', 'danger')
    else:
        flash('Turno no encontrado', 'warning')

    scroll_position = request.form.get('scroll_position', 0, type=int)
    return redirect(url_for('index', scroll_position=scroll_position))

@app.route('/borrar_turnos_seleccionados', methods=['POST'])
def borrar_turnos_seleccionados():
    if 'user_id' not in session:
        flash('Debe iniciar sesión primero', 'danger')
        return redirect(url_for('login'))

    scroll_position = request.form.get('scroll_position', 0, type=int)
    turno_ids = request.form.getlist('turno_ids')  # Get list of selected shift IDs

    if not turno_ids:
        flash('No se seleccionaron turnos para eliminar.', 'warning')
        return redirect(url_for('index', scroll_position=scroll_position))

    try:
        # Delete all shifts with the given IDs
        Turno.query.filter(Turno.id.in_(turno_ids)).delete()
        db.session.commit()
        flash(f'{len(turno_ids)} turno(s) eliminado(s) con éxito.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al eliminar los turnos: {str(e)}', 'danger')

    return redirect(url_for('index', scroll_position=scroll_position))

@app.route('/editar_turno/<int:id>', methods=['GET', 'POST'])
def editar_turno(id):
    if 'user_id' not in session:
        flash('Debe iniciar sesión primero', 'danger')
        return redirect(url_for('login'))

    turno = Turno.query.get(id)
    if not turno:
        flash('Turno no encontrado', 'danger')
        return redirect(url_for('index'))

    if request.method == 'POST':
        try:
            nombre = request.form.get('nombre')
            area = request.form.get('area')
            fecha_trabajo = request.form.get('fecha_trabajo')
            horas_trabajadas = int(request.form.get('horas_trabajadas', 6))
            turno_form = request.form.get('turno').upper()
            horas_normales = request.form.get('horas_normales')
            horas_extras = request.form.get('horas_extras')
            scroll_position = request.form.get('scroll_position', 0, type=int)

            if (horas_normales and horas_extras) or (not horas_normales and not horas_extras):
                flash("Debe seleccionar exactamente un tipo de horas: Normales o Extras.", "danger")
                return redirect(url_for("editar_turno", id=id, scroll_position=scroll_position))

            tipo_horas = "Normal" if horas_normales else "Extra"

            if turno_form == "NOCHE" and area not in opciones_nocturnas:
                flash("Área inválida para el turno Noche.", "danger")
                return redirect(url_for("editar_turno", id=id, scroll_position=scroll_position))
            elif turno_form != "NOCHE" and area not in opciones_diurnas:
                flash("Área inválida para el turno seleccionado.", "danger")
                return redirect(url_for("editar_turno", id=id, scroll_position=scroll_position))

            fecha = datetime.strptime(fecha_trabajo, "%Y-%m-%d").date()
            inicio_mes = date(fecha.year, fecha.month, 1)
            fin_mes = date(fecha.year, fecha.month, calendar.monthrange(fecha.year, fecha.month)[1])

            # Capitalizar el nombre
            nombre_capitalizado = capitalizar_nombre(nombre)

            # *** NEW VALIDATION: Check for existing shift with same date, turn, and area (excluding current shift) ***
            existing_turno_conflict = Turno.query.filter(
                Turno.fecha_trabajo == fecha,
                Turno.turno == turno_form,
                Turno.area == area,
                Turno.id != turno.id
            ).first()
            if existing_turno_conflict:
                flash(f"Ya existe un turno registrado para el área {area} en el turno {turno_form} del día {fecha.strftime('%d/%m/%Y')}.", "danger")
                return redirect(url_for("editar_turno", id=id, scroll_position=scroll_position))

            # Calculate current monthly hours (excluding current shift)
            total_horas_normales = db.session.query(func.sum(Turno.horas_trabajadas)).filter(
                and_(
                    func.lower(Turno.doctor) == nombre.lower(),
                    Turno.fecha_trabajo.between(inicio_mes, fin_mes),
                    Turno.tipo_horas == "Normal",
                    Turno.id != turno.id
                )
            ).scalar() or 0

            total_horas_extras = db.session.query(func.sum(Turno.horas_trabajadas)).filter(
                and_(
                    func.lower(Turno.doctor) == nombre.lower(),
                    Turno.fecha_trabajo.between(inicio_mes, fin_mes),
                    Turno.tipo_horas == "Extra",
                    Turno.id != turno.id
                )
            ).scalar() or 0

            if tipo_horas == "Normal" and total_horas_normales + horas_trabajadas > 150:
                flash(f"No se pueden asignar más de 150 horas normales por mes. Horas actuales: {total_horas_normales}", "danger")
                return redirect(url_for("editar_turno", id=id, scroll_position=scroll_position))

            if tipo_horas == "Extra" and total_horas_extras + horas_trabajadas > 100:
                flash(f"No se pueden asignar más de 100 horas extras por mes. Horas actuales: {total_horas_extras}", "danger")
                return redirect(url_for("editar_turno", id=id, scroll_position=scroll_position))

            # Update shift details
            turno.doctor = nombre_capitalizado
            turno.area = area
            turno.fecha_trabajo = fecha
            turno.horas_trabajadas = horas_trabajadas
            turno.turno = turno_form
            turno.tipo_horas = tipo_horas

            db.session.commit()
            flash('Turno actualizado correctamente', 'success')
            return redirect(url_for('index', scroll_position=scroll_position))
        except Exception as e:
            db.session.rollback()
            flash(f'Error al actualizar turno: {str(e)}', 'danger')
            return redirect(url_for("editar_turno", id=id, scroll_position=scroll_position))

    return render_template(
        'editar.html',
        turno=turno,
        opciones_diurnas=opciones_diurnas,
        opciones_nocturnas=opciones_nocturnas,
        datetime=datetime,
        doctores=Doctor.query.all(),
        scroll_position=request.args.get('scroll_position', 0, type=int)
    )

@app.route('/calendario')
def calendario():
    if 'user_id' not in session:
        flash('Debe iniciar sesión primero', 'danger')
        return redirect(url_for('login'))

    turnos = Turno.query.all()
    return render_template("calendario.html", turnos=turnos)

@app.route('/calendario/<int:mes>')
def calendario_mes(mes):
    if 'user_id' not in session:
        flash('Debe iniciar sesión primero', 'danger')
        return redirect(url_for('login'))

    if not (1 <= mes <= 12):
        flash("Mes inválido.", "danger")
        return redirect(url_for("index"))

    year = 2025  # Usamos 2025 explícitamente
    meses_nombres = [
        "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
        "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"
    ]
    nombre_mes = meses_nombres[mes - 1]
    dias_del_mes = calendar.monthrange(year, mes)[1]
    
    # Calcular el día de la semana del primer día (0=domingo, 6=sábado)
    primer_dia = datetime(year, mes, 1).date()
    primer_dia_semana = primer_dia.weekday()  # 0=lunes, 6=domingo
    primer_dia_semana = (primer_dia_semana + 1) % 7  # Ajustar: 0=domingo, 6=sábado

    inicio_mes = datetime(year, mes, 1).date()
    fin_mes = datetime(year, mes, dias_del_mes).date()

    # Obtener turnos del mes
    turnos = Turno.query.filter(
        Turno.fecha_trabajo >= inicio_mes,
        Turno.fecha_trabajo <= fin_mes
    ).all()

    # Generar colores para doctores
    doctores_unicos = list(set(turno.doctor for turno in turnos))
    def generar_colores_unicos(n):
        colores = []
        for i in range(n):
            h = i / n
            s = 0.7
            l = 0.5
            r, g, b = colorsys.hls_to_rgb(h, l, s)
            color_hex = f"#{int(r*255):02X}{int(g*255):02X}{int(b*255):02X}"
            colores.append(color_hex)
        return colores
    colores_generados = generar_colores_unicos(len(doctores_unicos))
    colores_doctores = {doctor: colores_generados[i] for i, doctor in enumerate(doctores_unicos)}

    # Generar semanas
    semanas = []
    current_date = inicio_mes
    if primer_dia_semana != 0:  # Si no empieza en domingo
        current_date -= timedelta(days=primer_dia_semana)
    
    # Generar hasta 6 semanas (máximo necesario para cualquier mes)
    for _ in range(6):
        semana = []
        for _ in range(7):  # Domingo a sábado
            if inicio_mes <= current_date <= fin_mes:
                dia_info = {'dia': current_date.day, 'turnos': [], 'fecha': current_date}
                for turno in turnos:
                    if turno.fecha_trabajo == current_date:
                        dia_info['turnos'].append(turno)
            else:
                dia_info = None  # Día vacío
            semana.append(dia_info)
            current_date += timedelta(days=1)
        if any(dia is not None for dia in semana):  # Solo agregar semanas con al menos un día del mes
            semanas.append(semana)
        else:
            break  # Salir si la semana no tiene días del mes

    return render_template(
        'calendario.html',
        turnos=turnos,
        nombre_mes=nombre_mes,
        year=year,
        dias_del_mes=dias_del_mes,
        primer_dia_semana=primer_dia_semana,
        colores_doctores=colores_doctores,
        datetime=datetime,
        ultimo_dia=dias_del_mes,
        semanas=semanas
    )

@app.route('/calendario_semanal/<int:mes>/<string:doctor>')
def calendario_semanal(mes, doctor):
    if 'user_id' not in session:
        flash('Debe iniciar sesión primero', 'danger')
        return redirect(url_for('login'))

    # Corrected condition: Check if mes is between 1 and 12
    if not (1 <= mes <= 12):
        flash("Mes inválido.", "danger")
        return redirect(url_for("index"))

    # Capitalizar el nombre del doctor
    doctor_capitalizado = capitalizar_nombre(doctor)
    doctor_record = Doctor.query.filter_by(name=doctor_capitalizado).first()
    if not doctor_record:
        flash("Doctor no encontrado.", "danger")
        return redirect(url_for("index"))

    nombres_meses = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
                     "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
    nombre_mes = nombres_meses[mes - 1]

    year = 2025  # Usamos 2025 explícitamente
    _, num_days = calendar.monthrange(year, mes)
    start_date = datetime(year, mes, 1).date()
    end_date = datetime(year, mes, num_days).date()

    # Obtener turnos del doctor para el mes
    turnos = Turno.query.filter(
        Turno.doctor == doctor_capitalizado,
        Turno.fecha_trabajo >= start_date,
        Turno.fecha_trabajo <= end_date
    ).all()

    # Organizar turnos por fecha y tipo
    dias = {}
    for turno in turnos:
        fecha = turno.fecha_trabajo
        turno_type = turno.turno.upper()
        if turno_type not in ["MAÑANA", "TARDE", "NOCHE"]:
            turno_type = "MAÑANA"
        if fecha not in dias:
            dias[fecha] = {'MAÑANA': [], 'TARDE': [], 'NOCHE': []}
        dias[fecha][turno_type].append({
            'area': turno.area,
            'horas': turno.horas_trabajadas,
            'tipo_horas': turno.tipo_horas
        })

    # Generar semanas completas
    semanas = []
    current_date = start_date
    # Ajustar el inicio de la primera semana al lunes anterior
    if current_date.weekday() != 0:  # 0 es lunes
        current_date -= timedelta(days=current_date.weekday())

    while current_date <= end_date or current_date.weekday() != 0:
        week_start = current_date
        week_end = week_start + timedelta(days=6)
        week_days = {}
        temp_date = week_start

        # Generar los 7 días de la semana
        for _ in range(7):
            if start_date <= temp_date <= end_date:
                # Día dentro del mes
                week_days[temp_date] = dias.get(temp_date, {'MAÑANA': [], 'TARDE': [], 'NOCHE': []})
            else:
                # Día fuera del mes (vacío)
                week_days[temp_date] = None
            temp_date += timedelta(days=1)

        if week_days:
            semanas.append({
                'start': week_start,
                'end': min(week_end, end_date),
                'dias': week_days
            })

        current_date = week_end + timedelta(days=1)

    # Colores para los doctores
    colores_doctores = {
        "Luz Mariuxi Murillo Calvache": "#D1E7DD",
        "Juan Perez": "#D1E7FF"
    }

    return render_template(
        'calendario_semanal.html',
        doctor=doctor_capitalizado,
        nombre_mes=nombre_mes,
        semanas=semanas,
        colores_doctores=colores_doctores
    )

@app.route('/reporte', methods=['GET'])
def reporte_mes():
    if 'user_id' not in session:
        flash('Debe iniciar sesión primero', 'danger')
        return redirect(url_for('login'))

    mes = request.args.get('mes', type=int)

    if not mes or mes < 1 or mes > 12:
        flash("Mes inválido", "danger")
        return redirect(url_for('index'))

    año_actual = datetime.now().year
    primer_dia = date(año_actual, mes, 1)
    ultimo_dia = date(año_actual, mes, calendar.monthrange(año_actual, mes)[1])

    turnos = Turno.query.filter(
        Turno.fecha_trabajo >= primer_dia,
        Turno.fecha_trabajo <= ultimo_dia
    ).all()

    templates_por_mes = {
        1: "reporte_enero.html",
        2: "reporte_febrero.html",
        3: "reporte_marzo.html",
        4: "reporte_abril.html",
        5: "reporte_mayo.html",
        6: "reporte_junio.html",
        7: "reporte_julio.html",
        8: "reporte_agosto.html",
        9: "reporte_septiembre.html",
        10: "reporte_octubre.html",
        11: "reporte_noviembre.html",
        12: "reporte_diciembre.html",
    }

    mes_nombre = [
        "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
        "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"
    ][mes-1]

    return render_template(
        templates_por_mes[mes],
        turnos=turnos,
        mes=mes,
        mes_nombre=mes_nombre,
        fecha_actual=datetime.now().strftime('%d/%m/%Y')
    )




@app.route('/descargar_excel')
def descargar_excel():
    if 'user_id' not in session:  # Verifica si el usuario está autenticado
        flash('Debe iniciar sesión primero', 'danger')  # Muestra mensaje de error si no está autenticado
        return redirect(url_for('login'))  # Redirige a la página de login

    try:
        mes = request.args.get('mes', default=datetime.now().month, type=int)  # Obtiene el mes de los parámetros
        año = request.args.get('año', default=datetime.now().year, type=int)  # Obtiene el año de los parámetros

        if not 1 <= mes <= 12:  # Valida que el mes esté entre 1 y 12
            flash("Mes inválido", "danger")  # Muestra mensaje de error si el mes es inválido
            return redirect(url_for('index'))  # Redirige a la página principal

        MESES = [
            "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
            "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"
        ]  # Lista de nombres de meses
        nombre_mes = MESES[mes - 1]  # Obtiene el nombre del mes según el índice

        primer_dia = date(año, mes, 1)  # Crea objeto de fecha para el primer día del mes
        ultimo_dia = date(año, mes, calendar.monthrange(año, mes)[1])  # Crea objeto de fecha para el último día del mes

        turnos = Turno.query.filter(
            Turno.fecha_trabajo >= primer_dia,
            Turno.fecha_trabajo <= ultimo_dia
        ).order_by(Turno.fecha_trabajo, Turno.area).all()  # Consulta los turnos del mes ordenados por fecha y área

        if not turnos and not Doctor.query.all():  # Verifica si no hay turnos ni doctores registrados
            flash(f"No hay turnos ni doctores registrados para {nombre_mes} {año}", "warning")  # Muestra mensaje de advertencia
            return redirect(url_for('index'))  # Redirige a la página principal

        def get_first_surname(full_name):  # Define función para extraer el primer apellido
            """Extrae el primer apellido de un nombre completo."""
            parts = full_name.strip().split()  # Divide el nombre en partes
            if len(parts) < 2:  # Si no hay apellido
                return full_name  # Devuelve el nombre completo
            return parts[1]  # Devuelve el primer apellido

        doctores_con_exceso = []  # Lista para doctores que exceden las horas permitidas
        horas_por_doctor = {}  # Diccionario para acumular horas por doctor
        for turno in turnos:  # Itera sobre los turnos
            doctor = turno.doctor  # Obtiene el doctor del turno
            if doctor not in horas_por_doctor:  # Inicializa el contador de horas si el doctor no está
                horas_por_doctor[doctor] = 0
            horas_por_doctor[doctor] += turno.horas_trabajadas  # Acumula las horas trabajadas
            if horas_por_doctor[doctor] > 250 and doctor not in doctores_con_exceso:  # Verifica si excede 250 horas
                doctores_con_exceso.append(doctor)  # Agrega el doctor a la lista de exceso
                flash(f"¡ATENCIÓN! El doctor {get_first_surname(doctor)} ha superado las 250 horas permitidas (total: {horas_por_doctor[doctor]} horas)", "danger")  # Muestra mensaje de advertencia

        wb = Workbook()  # Crea un nuevo libro de Excel
        ws = wb.active  # Selecciona la hoja activa
        ws.title = f"Turnos {nombre_mes}"  # Establece el título de la hoja
       

        header_font = Font(bold=True, color="FFFFFF")  # Define fuente para encabezados
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")  # Define relleno azul para encabezados
        pink_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Define relleno rosa
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Define relleno verde
        area_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")  # Define relleno para áreas
        center = Alignment(horizontal="center", vertical="center")  # Define alineación centrada
        wrap_text = Alignment(wrap_text=True, horizontal="center", vertical="center")  # Define alineación con ajuste de texto
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Define relleno amarillo
        light_blue_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")  # Define relleno azul claro
        total_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")  # Define relleno para totales

        border_style = Border(
            left=Side(style='medium', color='000000'),
            right=Side(style='medium', color='000000'),
            top=Side(style='medium', color='000000'),
            bottom=Side(style='medium', color='000000')
        )  # Define bordes medianos

        cell_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )  # Define bordes finos

        COLORES_DOCTORES = [
            "FF9999", "CC99FF", "99FF99", "99CCFF", "FFCC99",
            "FFFF99", "99FFFF", "CCFFCC", "FFCCFF", "FFCC66",
            "FF9966", "66CCCC", "CCCCCC", "FF99CC", "CCCC99",
            "FFCC00", "99CC99", "FFCCCC", "FFFFCC", "CCFFFF",
            "CC99CC", "99FFCC", "CCCCFF", "FF99FF", "99CCCC",
            "FF6666", "66FF99", "6699FF", "FFCC99", "99FF66",
            "CC6699", "66CC99", "9999FF", "FF99CC", "66CCCC"
        ]  # Lista de colores para doctores

        doctores_unicos = list({turno.doctor.lower() for turno in turnos})  # Obtiene doctores únicos de los turnos
        doctores_unicos.extend([doctor.name.lower() for doctor in Doctor.query.all()])  # Agrega doctores de la base de datos
        doctores_unicos = list(dict.fromkeys(doctores_unicos))  # Elimina duplicados
        doctor_colores = {doctor: COLORES_DOCTORES[i % len(COLORES_DOCTORES)]
                         for i, doctor in enumerate(doctores_unicos)}  # Asigna un color a cada doctor

        # Ajuste fijo de anchos y alturas para replicar mayo
        area_column_width = 20  # Ancho fijo para columnas de áreas (A y Q)
        surname_column_width = 12  # Ancho fijo para columnas de días
        area_row_height = 15  # Altura fija para filas de áreas
        header_row_height = 20  # Altura fija para filas de encabezados
        detail_row_height = 30  # Altura fija para filas de detalle
        summary_row_height = 30  # Altura fija para filas de resumen

        # Aplicar anchos de columna
        ws.column_dimensions['A'].width = area_column_width  # Columna A (áreas)
        ws.column_dimensions['Q'].width = area_column_width  # Columna Q (áreas adicionales)
        dias_mes = list(range(1, ultimo_dia.day + 1))  # Lista de días del mes
        for col, dia in enumerate(dias_mes, start=2):
            if dia <= 15:
                ws.column_dimensions[get_column_letter(col)].width = surname_column_width  # Columnas B-P (días 1-15)
            else:
                ws.column_dimensions[get_column_letter(col + 1)].width = surname_column_width  # Columnas R+ (días 16-31)

        ws.merge_cells('A1:B1')  # Combina celdas para el título del mes
        ws['A1'] = nombre_mes  # Establece el nombre del mes en la celda A1
        ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")  # Aplica fuente al título
        ws['A1'].alignment = center  # Centra el texto
        ws['A1'].fill = header_fill  # Aplica relleno azul
        ws['A1'].border = border_style  # Aplica borde mediano
        ws['B1'].border = border_style  # Aplica borde mediano a la celda B1
        ws.row_dimensions[1].height = header_row_height  # Altura fija para fila de título

        # Corregida para mayo de 2025: 1 de mayo es jueves
        DIAS_SEMANA = ["L", "M", "M", "J", "V", "S", "D"]  # Abreviaturas de días de la semana

        doctores_por_dia_area = {}  # Diccionario para agrupar doctores por turno, área y día
        for turno in turnos:  # Itera sobre los turnos
            if turno.turno == "NOCHE" and turno.area not in opciones_nocturnas:  # Para turnos nocturnos
                turno.area = "G.N."  # Asigna "G.N." si el área no es válida
            turno_tipo = (turno.turno or "MAÑANA").upper()  # Normaliza el tipo de turno
            if turno_tipo not in ["MAÑANA", "TARDE", "NOCHE"]:  # Verifica tipo de turno válido
                turno_tipo = "MAÑANA"  # Asigna "MAÑANA" por defecto
            key = (turno_tipo, turno.area, turno.fecha_trabajo.day)  # Crea clave para agrupar
            if key not in doctores_por_dia_area:  # Inicializa lista si la clave no existe
                doctores_por_dia_area[key] = []
            doctores_por_dia_area[key].append(turno)  # Agrega el turno a la lista

        # Configuración de turnos con áreas predefinidas
        turnos_config = [
            ("MAÑANA", pink_fill, opciones_diurnas),
            ("TARDE", pink_fill, opciones_diurnas),
            ("NOCHE", pink_fill, opciones_nocturnas)
        ]

        current_row = 3  # Inicia en la fila 3
        areas_rows = {}  # Diccionario para mapear filas de áreas

        for turno_idx, (turno_nombre, fill_color, areas_turno) in enumerate(turnos_config):  # Itera sobre configuración de turnos
            start_row = current_row  # Guarda la fila inicial del turno

            # Encabezado de turno en columna A (combinado sobre dos filas)
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row+1, end_column=1)  # Combina celdas para el nombre del turno
            ws.cell(row=current_row, column=1, value=turno_nombre)  # Establece el nombre del turno
            ws.cell(row=current_row, column=1).font = Font(bold=True, size=12)  # Aplica fuente
            ws.cell(row=current_row, column=1).alignment = center  # Centra el texto
            ws.cell(row=current_row, column=1).fill = fill_color  # Aplica relleno
            ws.cell(row=current_row, column=1).border = border_style  # Aplica borde

            # Encabezado de turno en columna Q (combinado sobre dos filas)
            split_col = 17  # Columna Q
            ws.merge_cells(start_row=current_row, start_column=split_col, end_row=current_row+1, end_column=split_col)  # Combina celdas para el nombre del turno
            ws.cell(row=current_row, column=split_col, value=turno_nombre)  # Establece el nombre del turno
            ws.cell(row=current_row, column=split_col).font = Font(bold=True, size=12)  # Aplica fuente
            ws.cell(row=current_row, column=split_col).alignment = center  # Centra el texto
            ws.cell(row=current_row, column=split_col).fill = fill_color  # Aplica relleno
            ws.cell(row=current_row, column=split_col).border = border_style  # Aplica borde

            ws.row_dimensions[current_row].height = header_row_height  # Altura fija para encabezado
            ws.row_dimensions[current_row+1].height = header_row_height  # Altura fija para encabezado

            for r in range(current_row, current_row+2):  # Aplica bordes a las filas del encabezado
                ws.cell(row=r, column=1).border = border_style
                ws.cell(row=r, column=split_col).border = border_style

            for col, dia_num in enumerate(dias_mes, start=2):  # Itera sobre los días del mes
                dia_letra = DIAS_SEMANA[date(año, mes, dia_num).weekday()]  # Obtiene la letra del día
                if dia_num <= 15:
                    col_adjusted = col  # Columnas B-P
                else:
                    col_adjusted = col + 1  # Columnas R+, desplazados por la nueva columna Q

                celda_dia = ws.cell(row=current_row, column=col_adjusted, value=dia_letra)  # Establece la letra del día
                celda_dia.fill = fill_color  # Aplica relleno
                celda_dia.alignment = center  # Centra el texto
                celda_dia.border = cell_border  # Aplica borde fino

                celda_num = ws.cell(row=current_row+1, column=col_adjusted, value=dia_num)  # Establece el número del día
                celda_num.fill = fill_color  # Aplica relleno
                celda_num.alignment = center  # Centra el texto
                celda_num.border = cell_border  # Aplica borde fino

            current_row += 2  # Avanza dos filas

            for area in areas_turno:  # Itera sobre las áreas del turno
                celda = ws.cell(row=current_row, column=1, value=area)  # Establece el nombre del área
                celda.alignment = wrap_text  # Aplica alineación con ajuste de texto
                celda.fill = area_fill  # Aplica relleno
                celda.border = border_style  # Aplica borde
                ws.row_dimensions[current_row].height = area_row_height  # Altura fija para áreas

                celda_split = ws.cell(row=current_row, column=split_col, value=area)  # Repite el nombre del área
                celda_split.alignment = wrap_text  # Aplica alineación con ajuste de texto
                celda_split.fill = area_fill  # Aplica relleno
                celda_split.border = border_style  # Aplica borde

                areas_rows[(turno_nombre, area)] = current_row  # Mapea la fila del área

                for c in range(1, len(dias_mes) + 3):  # Aplica bordes a la fila
                    ws.cell(row=current_row, column=c).border = cell_border

                current_row += 1  # Avengers una fila

            end_row = current_row - 1  # Actualiza la fila final

            for r in range(start_row, end_row + 1):  # Aplica bordes a los extremos
                ws.cell(row=r, column=1).border = border_style
                ws.cell(row=r, column=split_col).border = border_style
                ws.cell(row=r, column=len(dias_mes) + 2).border = border_style

            for c in range(1, len(dias_mes) + 3):  # Aplica bordes al encabezado y pie
                ws.cell(row=start_row, column=c).border = border_style
                ws.cell(row=end_row, column=c).border = border_style

            current_row += 4 if turno_nombre == "TARDE" else 3  # Avanza filas adicionales según turno

        turnos_por_tipo = {"MAÑANA": [], "TARDE": [], "NOCHE": []}  # Diccionario para turnos por tipo
        for turno in turnos:  # Itera sobre los turnos
            turno_tipo = (turno.turno or "MAÑANA").upper()  # Normaliza el tipo de turno
            if turno_tipo not in turnos_por_tipo:  # Verifica tipo válido
                turno_tipo = "MAÑANA"  # Asigna "MAÑANA" por defecto
            turnos_por_tipo[turno_tipo].append(turno)  # Agrega el turno al tipo correspondiente

        for key, turnos_dia in doctores_por_dia_area.items():  # Itera sobre turnos agrupados
            turno_tipo, area, dia = key  # Descompone la clave
            area_row = areas_rows.get((turno_tipo, area))  # Obtiene la fila del área
            if not area_row:  # Verifica si la fila existe
                continue
            if dia <= 15:
                dia_col = dia + 1  # Días 1-15 en columnas B-P
            else:
                dia_col = dia + 2  # Días 16+ en columnas R+
            doctores_info = []  # Lista para información de doctores
            for turno in turnos_dia:  # Itera sobre los turnos del día
                color_key = turno.doctor.lower()  # Obtiene clave del doctor
                color = doctor_colores.get(color_key, "FFFFFF")  # Obtiene color del doctor
                if turno.tipo_horas == "Extra":  # Verifica si es hora extra
                    color = green_fill.start_color  # Usa color verde
                full_doctor_name = next((d.name for d in Doctor.query.all() if d.name.lower() == turno.doctor.lower()), turno.doctor)  # Obtiene nombre completo
                surname = get_first_surname(full_doctor_name)  # Extrae apellido
                display_name = surname.upper() if turno.tipo_horas == "Extra" else surname.capitalize()  # Formatea nombre
                doctores_info.append({"nombre": display_name, "color": color, "tipo": turno.tipo_horas})  # Agrega información
            doctores_info.sort(key=lambda x: (x["tipo"] != "Normal", x["nombre"]))  # Ordena doctores
            texto_doctores = "\n".join(doc["nombre"] for doc in doctores_info)  # Combina nombres
            celda = ws.cell(row=area_row, column=dia_col, value=texto_doctores)  # Asigna texto a la celda
            celda.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")  # Aplica alineación con ajuste
            celda.border = cell_border  # Aplica borde
            celda.font = Font(color="000000", size=10, bold=True)  # Aplica fuente
            if len(doctores_info) > 1:  # Verifica si hay múltiples doctores
                colores = [doctor["color"] for doctor in doctores_info]  # Obtiene colores
                fill = GradientFill(stop=colores, type="linear", degree=90)  # Crea degradado
                celda.fill = fill  # Aplica degradado
            else:  # Caso de un solo doctor
                color = doctores_info[0]["color"] if doctores_info else "FFFFFF"  # Obtiene color
                celda.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")  # Aplica relleno sólido

        # Fijar alturas de filas para áreas
        for turno_nombre, _, areas_turno in turnos_config:
            for area in areas_turno:
                area_row = areas_rows.get((turno_nombre, area))
                if area_row:
                    ws.row_dimensions[area_row].height = area_row_height  # Altura fija para mayo

        ws_detalle = wb.create_sheet(title="Detalle Completo")  # Crea hoja de detalle
        headers = ["Doctor", "Área", "Fecha", "Horas Trabajadas", "Turno", "Tipo Horas"]  # Define encabezados
        for col, header in enumerate(headers, 1):  # Itera sobre encabezados
            celda = ws_detalle.cell(row=1, column=col, value=header)  # Establece encabezado
            celda.font = header_font  # Aplica fuente
            celda.fill = header_fill  # Aplica relleno
            celda.alignment = center  # Centra texto
            celda.border = border_style  # Aplica borde
        ws_detalle.row_dimensions[1].height = header_row_height  # Altura fija para encabezado

        for row, turno in enumerate(turnos, 2):  # Itera sobre turnos para llenar detalles
            color_key = turno.doctor.lower()  # Obtiene clave del doctor
            color_doctor = doctor_colores.get(color_key, "FFFFFF")  # Obtiene color
            full_doctor_name = next((d.name for d in Doctor.query.all() if d.name.lower() == turno.doctor.lower()), turno.doctor)  # Obtiene nombre completo
            celda_doctor = ws_detalle.cell(row=row, column=1, value=get_first_surname(full_doctor_name))  # Establece apellido
            celda_doctor.fill = PatternFill(
                start_color=color_doctor if turno.tipo_horas == "Normal" else green_fill.start_color,
                end_color=color_doctor if turno.tipo_horas == "Normal" else green_fill.end_color,
                fill_type="solid"
            )  # Aplica relleno según tipo de horas
            celda_doctor.font = Font(color="000000", size=12, bold=True)  # Aplica fuente
            celda_doctor.border = cell_border  # Aplica borde
            celda_doctor.alignment = wrap_text  # Aplica alineación con ajuste
            ws_detalle.cell(row=row, column=2, value=turno.area).border = cell_border  # Establece área
            ws_detalle.cell(row=row, column=2).alignment = wrap_text  # Aplica alineación
            ws_detalle.cell(row=row, column=3, value=turno.fecha_trabajo).number_format = 'DD/MM/YYYY'  # Establece fecha
            ws_detalle.cell(row=row, column=3).border = cell_border  # Aplica borde
            ws_detalle.cell(row=row, column=4, value=turno.horas_trabajadas).border = cell_border  # Establece horas
            ws_detalle.cell(row=row, column=4).alignment = center  # Centra texto
            ws_detalle.cell(row=row, column=5, value=turno.turno or "MAÑANA").border = cell_border  # Establece turno
            ws_detalle.cell(row=row, column=5).alignment = center  # Centra texto
            ws_detalle.cell(row=row, column=6, value=turno.tipo_horas).border = cell_border  # Establece tipo de horas
            ws_detalle.cell(row=row, column=6).alignment = center  # Centra texto
            ws_detalle.row_dimensions[row].height = detail_row_height  # Altura fija para detalle
            

        # Fijar anchos de columna para hoja de detalle
        ws_detalle.column_dimensions['A'].width = 20  # Ancho fijo para Doctor
        ws_detalle.column_dimensions['B'].width = 15  # Ancho fijo para Área
        ws_detalle.column_dimensions['C'].width = 12  # Ancho fijo para Fecha
        ws_detalle.column_dimensions['D'].width = 8   # Ancho fijo para Horas
        ws_detalle.column_dimensions['E'].width = 10  # Ancho fijo para Turno
        ws_detalle.column_dimensions['F'].width = 12  # Ancho fijo para Tipo Horas

        current_row += 5  # Avanza filas

        doctores_con_turnos = {turno.doctor.lower() for turno in turnos}  # Obtiene doctores únicos con turnos
        doctores_normalizados = {}
        for doctor in doctores_con_turnos:  # Itera sobre doctores con turnos
            full_name = next((d.name for d in Doctor.query.all() if d.name.lower() == doctor), doctor)  # Obtiene nombre completo
            doctores_normalizados[doctor] = full_name  # Normaliza el nombre
        doctores_min = sorted(doctores_normalizados.keys())  # Ordena doctores con turnos

        # Usar todas las áreas predefinidas para los resúmenes
        areas = opciones_diurnas + opciones_nocturnas  # Combinar todas las áreas para los resúmenes
        areas = sorted(list(set(areas)))  # Eliminar duplicados y ordenar

        ws.column_dimensions['A'].width = max(20, area_column_width)  # Actualiza ancho de columna A para doctores

        resumen_normal = {doctor: {area: 0 for area in areas} for doctor in doctores_min}  # Inicializa resumen de horas normales
        resumen_extra = {doctor: {area: 0 for area in areas} for doctor in doctores_min}  # Inicializa resumen de horas extra

        for turno in turnos:  # Itera sobre turnos para llenar resúmenes
            doctor_key = turno.doctor.lower()  # Obtiene clave del doctor
            area = turno.area if turno.area in areas else "G.N."  # Usa "G.N." si el área no está en la lista
            if turno.tipo_horas == "Normal":  # Verifica si es hora normal
                resumen_normal[doctor_key][area] += turno.horas_trabajadas  # Acumula horas normales
            else:  # Acumula horas extra
                resumen_extra[doctor_key][area] += turno.horas_trabajadas

        for doctor in doctores_min:  # Itera sobre doctores
            total_normal = sum(resumen_normal[doctor].values())  # Calcula total de horas normales
            if total_normal > 150:  # Verifica si excede 150 horas
                factor = 150 / total_normal  # Calcula factor de ajuste
                for area in resumen_normal[doctor]:  # Ajusta horas por área
                    resumen_normal[doctor][area] = int(resumen_normal[doctor][area] * factor)

            total_extra = sum(resumen_extra[doctor].values())  # Calcula total de horas extra
            if total_extra > 100:  # Verifica si excede 100 horas
                factor = 100 / total_extra  # Calcula factor de ajuste
                for area in resumen_extra[doctor]:  # Ajusta horas por área
                    resumen_extra[doctor][area] = int(resumen_extra[doctor][area] * factor)

        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(areas)+2)  # Combina celdas para título
        celda_titulo = ws.cell(row=current_row, column=1, value="HORAS NORMALES POR DOCTOR Y ÁREA")  # Establece título
        celda_titulo.font = Font(bold=True, size=12)  # Aplica fuente
        celda_titulo.fill = yellow_fill  # Aplica relleno
        celda_titulo.alignment = center  # Centra texto
        celda_titulo.border = border_style  # Aplica borde
        ws.row_dimensions[current_row].height = header_row_height  # Altura fija para título
        current_row += 1  # Avanza fila

        ws.cell(row=current_row, column=1, value="DOCTOR").font = Font(bold=True)  # Establece encabezado
        ws.cell(row=current_row, column=1).border = border_style  # Aplica borde
        ws.cell(row=current_row, column=1).alignment = center  # Centra texto
        for col, area in enumerate(areas, start=2):  # Itera sobre áreas
            ws.cell(row=current_row, column=col, value=area).font = Font(bold=True)  # Establece área
            ws.cell(row=current_row, column=col).fill = area_fill  # Aplica relleno
            ws.cell(row=current_row, column=col).border = border_style  # Aplica borde
            ws.cell(row=current_row, column=col).alignment = wrap_text  # Aplica alineación

        ws.cell(row=current_row, column=len(areas)+2, value="Total").font = Font(bold=True)  # Establece total
        ws.cell(row=current_row, column=len(areas)+2).fill = total_fill  # Aplica relleno
        ws.cell(row=current_row, column=len(areas)+2).alignment = center  # Centra texto
        ws.cell(row=current_row, column=len(areas)+2).border = border_style  # Aplica borde
        ws.row_dimensions[current_row].height = header_row_height  # Altura fija para encabezado
        current_row += 1  # Avanza fila

        for doctor in doctores_min:  # Itera sobre doctores con turnos
            total_doctor = 0  # Inicializa total
            nombre_mostrar = doctores_normalizados.get(doctor, doctor)  # Obtiene nombre normalizado
            celda_doctor = ws.cell(row=current_row, column=1, value=nombre_mostrar)  # Establece nombre
            celda_doctor.font = Font(color="000000", size=12, bold=True)  # Aplica fuente
            celda_doctor.border = border_style  # Aplica borde
            celda_doctor.alignment = wrap_text  # Aplica alineación

            for col, area in enumerate(areas, start=2):  # Itera sobre áreas
                horas = resumen_normal[doctor][area]  # Obtiene horas
                total_doctor += horas  # Acumula total
                celda = ws.cell(row=current_row, column=col, value=horas if horas > 0 else "")  # Establece horas
                celda.alignment = center  # Centra texto
                celda.border = cell_border  # Aplica borde
                if horas > 0:  # Verifica si hay horas
                    celda.fill = light_blue_fill  # Aplica relleno

            total_doctor = min(total_doctor, 150)  # Limita total a 150
            celda_total = ws.cell(row=current_row, column=len(areas)+2, value=total_doctor)  # Establece total
            celda_total.alignment = center  # Centra texto
            celda_total.fill = total_fill  # Aplica relleno
            celda_total.border = border_style  # Aplica borde
            celda_total.font = Font(bold=True)  # Aplica fuente
            ws.row_dimensions[current_row].height = summary_row_height  # Altura fija para resumen
            current_row += 1  # Avanza fila

        for c in range(1, len(areas)+3):  # Aplica bordes a encabezado y pie
            ws.cell(row=current_row - len(doctores_min) - 1, column=c).border = border_style
            ws.cell(row=current_row - 1, column=c).border = border_style

        for r in range(current_row - len(doctores_min) - 1, current_row):  # Aplica bordes a extremos
            ws.cell(row=r, column=1).border = border_style
            ws.cell(row=r, column=len(areas)+2).border = border_style

        current_row += 2  # Avanza filas

        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(areas)+2)  # Combina celdas para título
        celda_titulo = ws.cell(row=current_row, column=1, value="HORAS EXTRA POR DOCTOR Y ÁREA")  # Establece título
        celda_titulo.font = Font(bold=True, size=12)  # Aplica fuente
        celda_titulo.fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")  # Aplica relleno
        celda_titulo.alignment = center  # Centra texto
        celda_titulo.border = border_style  # Aplica borde
        ws.row_dimensions[current_row].height = header_row_height  # Altura fija para título
        current_row += 1  # Avanza fila

        ws.cell(row=current_row, column=1, value="DOCTOR").font = Font(bold=True)  # Establece encabezado
        ws.cell(row=current_row, column=1).border = border_style  # Aplica borde
        ws.cell(row=current_row, column=1).alignment = center  # Centra texto
        for col, area in enumerate(areas, start=2):  # Itera sobre áreas
            ws.cell(row=current_row, column=col, value=area).font = Font(bold=True)  # Establece área
            ws.cell(row=current_row, column=col).fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")  # Aplica relleno
            ws.cell(row=current_row, column=col).border = border_style  # Aplica borde
            ws.cell(row=current_row, column=col).alignment = wrap_text  # Aplica alineación

        ws.cell(row=current_row, column=len(areas)+2, value="Total").font = Font(bold=True)  # Establece total
        ws.cell(row=current_row, column=len(areas)+2).fill = total_fill  # Aplica relleno
        ws.cell(row=current_row, column=len(areas)+2).alignment = center  # Centra texto
        ws.cell(row=current_row, column=len(areas)+2).border = border_style  # Aplica borde
        ws.row_dimensions[current_row].height = header_row_height  # Altura fija para encabezado
        current_row += 1  # Avanza fila

        for doctor in doctores_min:  # Itera sobre doctores con turnos
            total_doctor = 0  # Inicializa total
            nombre_mostrar = doctores_normalizados.get(doctor, doctor)  # Obtiene nombre normalizado
            celda_doctor = ws.cell(row=current_row, column=1, value=nombre_mostrar)  # Establece nombre
            celda_doctor.font = Font(color="000000", size=12, bold=True)  # Aplica fuente
            celda_doctor.border = border_style  # Aplica borde
            celda_doctor.alignment = wrap_text  # Aplica alineación

            for col, area in enumerate(areas, start=2):  # Itera sobre áreas
                horas = resumen_extra[doctor][area]  # Obtiene horas
                total_doctor += horas  # Acumula total
                celda = ws.cell(row=current_row, column=col, value=horas if horas > 0 else "")  # Establece horas
                celda.alignment = center  # Centra texto
                celda.border = cell_border  # Aplica borde
                if horas > 0:  # Verifica si hay horas
                    celda.fill = PatternFill(start_color="FFDDDD", end_color="FFDDDD", fill_type="solid")  # Aplica relleno

            total_doctor = min(total_doctor, 100)  # Limita total a 100
            celda_total = ws.cell(row=current_row, column=len(areas)+2, value=total_doctor)  # Establece total
            celda_total.alignment = center  # Centra texto
            celda_total.fill = total_fill  # Aplica relleno
            celda_total.border = border_style  # Aplica borde
            celda_total.font = Font(bold=True)  # Aplica fuente
            ws.row_dimensions[current_row].height = summary_row_height  # Altura fija para resumen
            current_row += 1  # Avanza fila

        for c in range(1, len(areas)+3):  # Aplica bordes a encabezado y pie
            ws.cell(row=current_row - len(doctores_min) - 1, column=c).border = border_style
            ws.cell(row=current_row - 1, column=c).border = border_style

        for r in range(current_row - len(doctores_min) - 1, current_row):  # Aplica bordes a extremos
            ws.cell(row=r, column=1).border = border_style
            ws.cell(row=r, column=len(areas)+2).border = border_style

        if doctores_con_exceso:  # Verifica si hay doctores con exceso
            current_row += 3  # Avanza filas
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(dias_mes)+2)  # Combina celdas
            celda_advertencia = ws.cell(row=current_row, column=1,
                                        value=f"ADVERTENCIA: Los siguientes doctores han excedido 250 horas: {', '.join(get_first_surname(doc) for doc in doctores_con_exceso)}")  # Establece advertencia
            celda_advertencia.font = Font(bold=True, color="FF0000")  # Aplica fuente
            celda_advertencia.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Aplica relleno
            celda_advertencia.alignment = center  # Centra texto
            ws.row_dimensions[current_row].height = header_row_height  # Altura fija para advertencia

        # Ajuste final de alturas para filas específicas
        ws.row_dimensions[48].height = 45  # Altura para la fila 48
        ws.row_dimensions[54].height = 45  # Altura para la fila 54

        output = BytesIO()  # Crea buffer para el archivo
        wb.save(output)  # Guarda el libro en el buffer
        output.seek(0)  # Posiciona el cursor al inicio

        nombre_archivo = f"Turnos_{nombre_mes}_{año}.xlsx"  # Define nombre del archivo

        return send_file(
            output,
            as_attachment=True,
            download_name=nombre_archivo,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )  # Envía el archivo como descarga

    except Exception as e:  # Captura cualquier error
        flash(f"Error al generar el reporte: {str(e)}", "danger")  # Muestra mensaje de error
        return redirect(url_for('index'))  # Redirige a la página principal

    
if __name__ == '__main__':
    with app.app_context():
        db.create_all()
        
        # Normalizar nombres en la tabla Doctor
        doctores = Doctor.query.all()
        for doctor in doctores:
            doctor.name = capitalizar_nombre(doctor.name)
        
        # Normalizar nombres en la tabla Turno
        turnos = Turno.query.all()
        for turno in turnos:
            turno.doctor = capitalizar_nombre(turno.doctor)
        
        try:
            db.session.commit()
            print("Nombres normalizados correctamente.")
        except Exception as e:
            db.session.rollback()
            print(f"Error al normalizar nombres: {str(e)}")
    
    app.run(debug=True)